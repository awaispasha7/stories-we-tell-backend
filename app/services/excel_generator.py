"""
Excel Generator Service
Generates Excel files from dossier data for email attachments
"""

import os
import tempfile
from typing import Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def generate_dossier_excel(dossier_data: Dict[str, Any], project_id: str) -> Optional[str]:
    """
    Generate an Excel file from dossier data.
    
    Args:
        dossier_data: Dictionary containing dossier information
        project_id: Project ID for filename
    
    Returns:
        Path to the generated Excel file, or None if generation fails
    """
    try:
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Story Dossier"
        
        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        label_font = Font(bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        row = 1
        
        # Title
        ws.merge_cells(f'A{row}:B{row}')
        title_cell = ws[f'A{row}']
        title_cell.value = f"Story Dossier - {dossier_data.get('title', 'Untitled Story')}"
        title_cell.font = Font(bold=True, size=14)
        title_cell.fill = header_fill
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 2
        
        # Project ID
        ws[f'A{row}'] = "Project ID:"
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'] = project_id
        row += 1
        
        # Story Overview Section
        row = _add_section_header(ws, row, "Story Overview", header_fill, header_font, border)
        
        if dossier_data.get('title'):
            row = _add_field(ws, row, "Title", dossier_data['title'], label_font, border)
        if dossier_data.get('logline'):
            row = _add_field(ws, row, "Logline", dossier_data['logline'], label_font, border)
        if dossier_data.get('genre'):
            row = _add_field(ws, row, "Genre", dossier_data['genre'], label_font, border)
        if dossier_data.get('tone'):
            row = _add_field(ws, row, "Tone", dossier_data['tone'], label_font, border)
        
        row += 1
        
        # Hero Characters Section
        heroes = dossier_data.get('heroes', [])
        if heroes:
            row = _add_section_header(ws, row, "Hero Characters", header_fill, header_font, border)
            for idx, hero in enumerate(heroes, 1):
                row = _add_section_header(ws, row, f"Hero {idx}", PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"), Font(bold=True, size=11), border)
                if hero.get('name'):
                    row = _add_field(ws, row, "Name", hero['name'], label_font, border)
                if hero.get('age_at_story'):
                    row = _add_field(ws, row, "Age at Story", hero['age_at_story'], label_font, border)
                if hero.get('relationship_to_user'):
                    row = _add_field(ws, row, "Relationship to User", hero['relationship_to_user'], label_font, border)
                if hero.get('physical_descriptors'):
                    row = _add_field(ws, row, "Physical Descriptors", hero['physical_descriptors'], label_font, border)
                if hero.get('personality_traits'):
                    row = _add_field(ws, row, "Personality Traits", hero['personality_traits'], label_font, border)
                if hero.get('photo_url'):
                    row = _add_field(ws, row, "Photo URL", hero['photo_url'], label_font, border)
                row += 1
        
        # Supporting Characters Section
        supporting = dossier_data.get('supporting_characters', [])
        if supporting:
            row = _add_section_header(ws, row, "Supporting Characters", header_fill, header_font, border)
            for idx, char in enumerate(supporting, 1):
                row = _add_section_header(ws, row, f"Supporting Character {idx}", PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"), Font(bold=True, size=11), border)
                if char.get('name'):
                    row = _add_field(ws, row, "Name", char['name'], label_font, border)
                if char.get('role'):
                    row = _add_field(ws, row, "Role", char['role'], label_font, border)
                if char.get('description'):
                    row = _add_field(ws, row, "Description", char['description'], label_font, border)
                if char.get('photo_url'):
                    row = _add_field(ws, row, "Photo URL", char['photo_url'], label_font, border)
                row += 1
        
        # Setting & Time Section
        if dossier_data.get('story_location') or dossier_data.get('story_timeframe') or dossier_data.get('season_time_of_year') or dossier_data.get('environmental_details'):
            row = _add_section_header(ws, row, "Setting & Time", header_fill, header_font, border)
            if dossier_data.get('story_location'):
                row = _add_field(ws, row, "Location", dossier_data['story_location'], label_font, border)
            if dossier_data.get('story_timeframe'):
                row = _add_field(ws, row, "Timeframe", dossier_data['story_timeframe'], label_font, border)
            if dossier_data.get('season_time_of_year'):
                row = _add_field(ws, row, "Season/Time of Year", dossier_data['season_time_of_year'], label_font, border)
            if dossier_data.get('environmental_details'):
                row = _add_field(ws, row, "Environmental Details", dossier_data['environmental_details'], label_font, border)
            row += 1
        
        # Story Type & Perspective Section
        if dossier_data.get('story_type') or dossier_data.get('perspective') or dossier_data.get('audience'):
            row = _add_section_header(ws, row, "Story Type & Perspective", header_fill, header_font, border)
            if dossier_data.get('story_type'):
                row = _add_field(ws, row, "Story Type", dossier_data['story_type'].replace('_', ' ').title(), label_font, border)
            audience = dossier_data.get('audience', {})
            if isinstance(audience, dict):
                if audience.get('who_will_see_first'):
                    row = _add_field(ws, row, "Audience", audience['who_will_see_first'], label_font, border)
                if audience.get('desired_feeling'):
                    row = _add_field(ws, row, "Desired Feeling", audience['desired_feeling'], label_font, border)
            if dossier_data.get('perspective'):
                row = _add_field(ws, row, "Perspective", dossier_data['perspective'].replace('_', ' ').title(), label_font, border)
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 60
        
        # Save to temporary file
        temp_dir = tempfile.gettempdir()
        excel_filename = f"dossier_{project_id}.xlsx"
        excel_path = os.path.join(temp_dir, excel_filename)
        wb.save(excel_path)
        
        print(f"✅ [EXCEL] Generated Excel file: {excel_path}")
        return excel_path
        
    except Exception as e:
        print(f"❌ [EXCEL] Error generating Excel file: {e}")
        import traceback
        print(f"❌ [EXCEL] Traceback: {traceback.format_exc()}")
        return None


def _add_section_header(ws, row: int, title: str, fill: PatternFill, font: Font, border: Border) -> int:
    """Add a section header row"""
    ws.merge_cells(f'A{row}:B{row}')
    cell = ws[f'A{row}']
    cell.value = title
    cell.font = font
    cell.fill = fill
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.border = border
    ws.row_dimensions[row].height = 25
    return row + 1


def _add_field(ws, row: int, label: str, value: Any, label_font: Font, border: Border) -> int:
    """Add a field row (label and value)"""
    label_cell = ws[f'A{row}']
    label_cell.value = label
    label_cell.font = label_font
    label_cell.border = border
    label_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    value_cell = ws[f'B{row}']
    value_cell.value = str(value) if value is not None else ""
    value_cell.border = border
    value_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    ws.row_dimensions[row].height = 20
    return row + 1

